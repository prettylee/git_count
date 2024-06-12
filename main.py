import logging
import openpyxl
import os
import subprocess
import time
from configparser import ConfigParser


# 获取配置文件
cf = ConfigParser()
cf.read('config.ini', encoding='utf-8')
# 获取git的bash，在windows系统需要获取，linux不需要这部
git_exe = cf.get('git', 'GitExe')


# 设置日志输出
def set_log():
    logger = logging.getLogger()
    logger.setLevel('INFO')
    self_myfmt = '%(asctime)s - %(funcName)s - %(levelname)s - %(message)s'
    formatter = logging.Formatter(self_myfmt)
    control_handler = logging.StreamHandler()
    control_handler.setLevel('INFO')
    control_handler.setFormatter(formatter)
    logger.addHandler(control_handler)


def test():
    logging.info("你好")
    ext = cf.get('ext', 'Ext')
    print(ext)


def read_xlsx(file_name, sheet_name):
    """
    :param file_name: xlsx文件名称
    :param sheet_name: sheet名称
    :return: 读取的内容
    """
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.get_sheet_by_name(sheet_name)
    data = list()
    for row in sheet.rows:
        temp_dict = dict()
        code = row[0].value
        branch = row[1].value
        ext = row[2].value
        if code is not None:
            temp_dict['code'] = code
            temp_dict['branch'] = branch
            temp_dict['ext'] = ext
        else:
            break

        data.append(temp_dict)
    # logging.info(data)
    data.pop(0)
    return data


def format_code_url(code_url):
    """
    将代码地址截取出项目名
    :param code_url:
    :return: project_name 项目目录项目名
    """
    project_name = code_url.split("/")[-1].split(".")[0]
    # logging.info("当前project name是：{}".format(project_name))
    return project_name


def clone_code(code_url, branch):
    """
    检验并clone代码
    :param code_url: git@codeup.aliyun.com:lonch/YYF/cloudpharmacy-drug-safety.git
    :param branch: master
    :return:
    """
    logging.info("——————开始检验代码库是否存在——————")
    code_path = '{}{}'.format(cf.get('pwd', 'pwd'), format_code_url(code_url))
    if os.path.exists(code_path):
        s = subprocess.Popen('git rev-parse --is-inside-work-tree',
                             shell=True,
                             cwd=code_path,
                             stdout=subprocess.PIPE).stdout
        s_output = s.read().decode('utf-8').strip()
        if s_output == 'true':
            logging.info('当前代码库是：{}'.format(format_code_url(code_url)))
            logging.info('-----开始进行pull操作拉去最新代码-----')
            git_pull_s = subprocess.Popen('git pull',
                                          shell=True,
                                          cwd=code_path,
                                          stdout=subprocess.PIPE).stdout
            logging.info(git_pull_s.read().decode('utf-8').strip())
            logging.info('-----git pull 操作执行完成-----')
        else:
            logging.warning('{}非代码库目录'.format(code_path))
            del_dir_s = subprocess.Popen(['{}'.format(git_exe), '-c', 'rm -rf {}'.format(code_path)],
                                         shell=False, stderr=subprocess.PIPE).stdout
            logging.info('{}'.format(del_dir_s.read().decode('utf-8').strip()))
            logging.info('开始克隆---->{}'.format(code_path))
            p2 = subprocess.Popen('git clone {} --branch {}'.format(code_url, branch),
                                  shell=True,
                                  cwd=cf.get('pwd', 'pwd'),
                                  stdout=subprocess.PIPE).stdout
            logging.info(p2.read().decode('utf-8').strip())
    else:
        p2 = subprocess.Popen('git clone {} --branch {}'.format(code_url, branch),
                              shell=True,
                              cwd=cf.get('pwd', 'pwd'),
                              stdout=subprocess.PIPE).stdout
        logging.info(p2.read().decode('utf-8').strip())


def code_count(code_url, ext):
    """
    代码的全量统计
    :param code_url: 代码url
    :param ext: 需要统计代码文件的扩展名
    :return: []
    """
    code_path = format_code_url(code_url)
    data = list()
    dead_line = cf.get('time', 'DeadLine')
    sub = subprocess.Popen(['{}'.format(git_exe), '-c',
                            "git log --until='%s' --pretty=tformat: --numstat | grep -E '%s' | awk "
                            "'{add += $1; subs += $2; loc += $1 - $2} END { printf loc }' " % (dead_line, ext)],
                            shell=False, cwd=code_path, stdout=subprocess.PIPE).stdout
    data.append(code_path)
    data.append(sub.read().decode('GBK').strip())
    return data


def commit_date_count(code_url):
    """
    统计提交次数和最后提交时间
    :param code_url: git仓库地址
    :return: ['42', '2022-06-30 23:59:59']
    """
    data = list()
    code_path = format_code_url(code_url)
    start_time = cf.get('time', 'StartLine')
    dead_line = cf.get('time', 'DeadLine')
    sub_commit_date = subprocess.Popen(['{}'.format(git_exe), '-c',
                                        "git log -n1  --until='%s' --date=format:'%%Y-%%m-%%d %%H:%%M:%%S' --no-merges | grep"
                                        " Date | awk '{print $2,$3}'" % (dead_line)],
                                       shell=False, cwd=code_path, stdout=subprocess.PIPE).stdout
    sub_commit_count = subprocess.Popen(['{}'.format(git_exe), '-c',
                                         "git log --since='%s' --until='%s' --no-merges |grep "
                                         "-e 'commit [a-zA-Z0-9]*' | wc -l" % (start_time, dead_line)],
                                        shell=False, cwd=code_path,
                                        stdout=subprocess.PIPE).stdout
    data.append(sub_commit_count.read().decode('GBK').strip())
    data.append(sub_commit_date.read().decode('GBK').strip())
    return data


def write_xlsx(data_temp, file_name):
    """
    写入excel
    :param data_temp: 数据列表
    :param file_name: 文件名称
    :return:
    """
    time.sleep(2)
    localtime = (time.strftime('%Y-%m-%d %H.%M.%S', time.localtime()))
    wb = openpyxl.load_workbook(file_name)
    ws1 = wb.create_sheet()
    ws1.title = '{}'.format(localtime)

    for row in data_temp:
        ws1.append(row)
    try:
        wb.save(file_name)
    except Exception as e:
        print(e)


def code_count_by_author(code_url, ext):
    """
    统计本月个人代码提交量
    :param code_url:
    :param ext:
    :return:
    """
    code_path = format_code_url(code_url)
    # print(code_path)
    # print(os.getcwd())
    start_line = cf.get('time', 'StartLine')
    dead_line = cf.get('time', 'DeadLine')
    scrpits = """
        git log --format='%aN' | sort -u | while read name; do echo -en "$name,"; done
        """
    r = None
    try:
        sub = subprocess.Popen(['{}'.format(git_exe), '-c', scrpits],
                               shell=False, cwd=code_path,
                               stdout=subprocess.PIPE).stdout
        r = sub.read().decode('utf-8')
    except Exception as e:
        logging.error("出现异常：{}".format(e))
    logging.info("====仓库：{} # ====name:{}".format(code_path, r))
    names = r.split(',')
    names.pop()
    logging.info("格式化name：{}".format(names))

    code_count_data = list()
    for name in names:
        temp_list = []
        shell = "git log --author='{}' --since='{}'  --until='{}' --pretty=tformat: " \
                "--numstat | grep -E '{}' | ".format(name, start_line, dead_line, ext)

        scrpits2 = shell + """awk '{add += $1; subs += $2; loc += $1 - $2} END { printf "%s,%s,%s",add,subs,loc }'"""

        logging.info('执行命令：{}'.format(scrpits2))
        sub2 = subprocess.Popen(['{}'.format(git_exe), '-c', scrpits2],
                                shell=False, cwd=code_path,
                                stdout=subprocess.PIPE).stdout
        r2 = sub2.read().decode('utf-8')
        logging.info('r2: {}'.format(r2))
        if r2 != ",,":
            logging.info("当前name：{} | 代码量：{}".format(name, r2))
            temp_list.append(name)
            temp_list = temp_list + r2.split(',')
            code_count_data.append(temp_list)

    logging.info("data: {}".format(code_count_data))
    return code_count_data


def run():
    set_log()
    os.chdir(cf.get('pwd', 'pwd'))
    file = 'C:\\Users\\Administrator\\Desktop\\git_count\\code.xlsx'
    git_data = read_xlsx(file, 'test')
    data_temp = [['项目', '行数', '本月提交次数', '最后提交时间']]
    head_temp = [['UserName', '增加的行数', '删除的行数', '增长量']]
    ext = cf.get('ext', 'Ext')
    for row in git_data:
        logging.info(row)
        code_url = row['code']
        branch = row['branch']
        # 克隆代码
        clone_code(code_url, branch)

        # 统计代码总量
        code_count_data_list = code_count(code_url, ext)

        # 统计提交次数和最后提交时间
        commit_date_count_list = commit_date_count(code_url)

        # 将两个得到的数据合并
        code_count_list = code_count_data_list + commit_date_count_list

        # 将表格头合并
        data_temp.append(code_count_list)

        # 统计本月个人代码提交
        author_data = code_count_by_author(code_url, ext)
        for i in author_data:
            head_temp.append(i)

    # 总量统计写入表格
    write_xlsx(data_temp, file)
    write_xlsx(head_temp, file)


if __name__ == "__main__":
    run()

